from fastapi import FastAPI, HTTPException, Header, UploadFile, File, BackgroundTasks, Request, Query
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, field_validator
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from apscheduler.schedulers.background import BackgroundScheduler
import asyncio
import os
import sys
import json
import time
import hashlib
import tarfile
import glob
import io
from urllib.parse import quote
from datetime import datetime

sys.path.append(os.path.dirname(__file__))

from contextlib import asynccontextmanager
from qa_engine import QAEngine
from logger import Logger
from pdf_monitor import PDFMonitor, PDFWatcher
from vector_store import VectorStore
from pdf_processor import PDFProcessor, load_all_pdfs_from_folder
from calc_engine import calculate_all, VALID_SPEEDS

# ─── 전역 상태 ───────────────────────────────────────────────

qa_engine = None
pdf_watcher = None
scheduler = None
rebuild_status = {"running": False, "message": "대기 중"}

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "roadspec2024")

# 응답 캐시 {hash: {answer, sources, time}}
_cache = {}
_cache_hits = 0
CACHE_TTL = 3600  # 1시간

# ─── 헬퍼 함수 ───────────────────────────────────────────────

def _check_admin(password: str):
    if not password or password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="비밀번호가 틀렸습니다.")

def _cache_key(question: str) -> str:
    return hashlib.md5(question.strip().lower().encode()).hexdigest()

def _get_cached(question: str):
    key = _cache_key(question)
    if key in _cache:
        entry = _cache[key]
        if time.time() - entry["time"] < CACHE_TTL:
            return entry
        del _cache[key]
    return None

def _set_cache(question: str, answer: str, sources: list):
    _cache[_cache_key(question)] = {"answer": answer, "sources": sources, "time": time.time()}

def _backup_vectordb():
    """vectordb 압축 백업 (최근 5개 유지)"""
    if not os.path.exists("vectordb"):
        return
    os.makedirs("backups", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join("backups", f"vectordb_{timestamp}.tar.gz")
    with tarfile.open(backup_path, "w:gz") as tar:
        tar.add("vectordb", arcname="vectordb")
    # 최근 5개만 유지
    backups = sorted(glob.glob(os.path.join("backups", "vectordb_*.tar.gz")))
    for old in backups[:-5]:
        os.remove(old)
    print(f"✓ vectordb 백업 완료: {backup_path}")

def _handle_new_pdfs(filenames):
    """PDF 추가/변경 시 벡터 DB 업데이트 + BM25 재구축"""
    vector_store = qa_engine.vector_store
    monitor = PDFMonitor()

    new_chunks = []
    for filename in filenames:
        filepath = os.path.join("data", filename)
        if not os.path.exists(filepath):
            continue
        processor = PDFProcessor(filepath)
        chunks = vector_store.create_chunks_from_pdf(processor)
        new_chunks.extend(chunks)
        print(f"  ✓ {filename}: {len(chunks)}개 청크 생성")

    if new_chunks:
        vector_store.update_vectorstore(new_chunks)
        qa_engine._build_bm25_index()
        print(f"✓ 벡터 DB + BM25 인덱스 업데이트 완료! ({len(new_chunks)}개 청크 추가)")

    monitor.update_state()

# ─── 앱 초기화 ───────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app):
    global qa_engine, pdf_watcher, scheduler

    print("QA 엔진 초기화 중...")
    qa_engine = QAEngine()
    print("✓ QA 엔진 초기화 완료!")

    pdf_watcher = PDFWatcher(data_folder="data", on_update=_handle_new_pdfs)
    pdf_watcher.start()

    # 스케줄러: 매일 새벽 3시 백업 + 로그 정리
    scheduler = BackgroundScheduler()
    scheduler.add_job(_backup_vectordb, 'cron', hour=3, minute=0)
    scheduler.add_job(lambda: Logger.cleanup_old_logs(days=30), 'cron', hour=3, minute=30)
    scheduler.start()
    print("✓ 스케줄러 시작 (매일 03:00 백업, 03:30 로그 정리)")

    yield

    pdf_watcher.stop()
    scheduler.shutdown()

limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="Roadspec API",
    description="도로설계 기준 문서 기반 질의응답 API",
    version="1.0.0",
    lifespan=lifespan
)

app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── 모델 ────────────────────────────────────────────────────

class ChatRequest(BaseModel):
    question: str
    session_id: str
    history: list = []  # [{"role": "user"/"assistant", "content": "..."}]
    selected_sources: list = []  # 빈 배열 = 전체 검색

    @field_validator('question')
    @classmethod
    def check_length(cls, v):
        v = v.strip()
        if len(v) == 0:
            raise ValueError('질문을 입력해주세요.')
        if len(v) > 500:
            raise ValueError('질문은 500자 이내로 입력해주세요.')
        return v

class ChatResponse(BaseModel):
    answer: str
    sources: list
    session_id: str

class FeedbackRequest(BaseModel):
    session_id: str
    question: str
    rating: int  # 1 = 좋아요, -1 = 싫어요

# ─── 채팅 엔드포인트 ─────────────────────────────────────────

@app.post("/api/chat/stream")
@limiter.limit("20/minute")
async def chat_stream(request: Request, body: ChatRequest):
    # 캐시 확인 (히스토리 없을 때만, 문서 필터 없을 때만)
    if not body.history and not body.selected_sources:
        cached = _get_cached(body.question)
        if cached:
            global _cache_hits
            _cache_hits += 1
            async def cached_gen():
                yield f"data: {json.dumps({'type': 'cache_hit'})}\n\n"
                for token in cached["answer"]:
                    yield f"data: {json.dumps({'type': 'token', 'content': token})}\n\n"
                yield f"data: {json.dumps({'type': 'sources', 'sources': cached['sources']})}\n\n"
                yield "data: [DONE]\n\n"
            return StreamingResponse(
                cached_gen(),
                media_type="text/event-stream",
                headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
            )

    # 검색은 스레드풀에서 실행 (BM25 CPU 블로킹 방지)
    loop = asyncio.get_event_loop()
    filter_sources = body.selected_sources or None
    stream, sources = await loop.run_in_executor(
        None, lambda: qa_engine.ask_stream(
            body.question,
            history=body.history,
            filter_sources=filter_sources
        )
    )

    async def generate():
        full_answer = ""
        try:
            for chunk in stream:
                token = chunk.choices[0].delta.content
                if token:
                    full_answer += token
                    yield f"data: {json.dumps({'type': 'token', 'content': token})}\n\n"

            if not body.history and not body.selected_sources:
                _set_cache(body.question, full_answer, sources)

            logger = Logger(session_id=body.session_id)
            logger.log(question=body.question, answer=full_answer, sources=sources)

            yield f"data: {json.dumps({'type': 'sources', 'sources': sources})}\n\n"

            # 연관 질문 추천 (히스토리·필터 없을 때만)
            if not body.history and not body.selected_sources:
                related = qa_engine.get_related_questions(body.question)
                if related:
                    yield f"data: {json.dumps({'type': 'related_questions', 'questions': related})}\n\n"

            yield "data: [DONE]\n\n"
        except Exception as e:
            try:
                Logger(session_id=body.session_id).log_error(
                    question=body.question,
                    error=str(e),
                    error_type=type(e).__name__,
                )
            except Exception:
                pass
            yield f"data: {json.dumps({'type': 'error', 'content': str(e)})}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )

# ─── 피드백 엔드포인트 ───────────────────────────────────────

@app.post("/api/feedback")
async def submit_feedback(body: FeedbackRequest):
    logger = Logger(session_id=body.session_id)
    logger.log_feedback(body.question, body.rating)
    return {"message": "피드백 감사합니다"}

@app.get("/api/admin/feedback")
async def admin_feedback_stats(x_admin_password: str = Header(None)):
    _check_admin(x_admin_password)
    feedback_data = []
    log_dir = "logs"
    if os.path.exists(log_dir):
        for session in os.listdir(log_dir):
            session_dir = os.path.join(log_dir, session)
            if not os.path.isdir(session_dir):
                continue
            for filename in os.listdir(session_dir):
                if not (filename.startswith("feedback_") and filename.endswith(".json")):
                    continue
                try:
                    with open(os.path.join(session_dir, filename), 'r', encoding='utf-8') as f:
                        feedbacks = json.load(f)
                    feedback_data.extend(feedbacks)
                except Exception:
                    pass
    stats = {
        "total": len(feedback_data),
        "positive": sum(1 for f in feedback_data if f.get("rating") == 1),
        "negative": sum(1 for f in feedback_data if f.get("rating") == -1),
    }
    recent = sorted(feedback_data, key=lambda x: x.get("timestamp", ""), reverse=True)[:50]
    return {"feedbacks": recent, "stats": stats}

@app.get("/api/admin/errors")
async def admin_errors(x_admin_password: str = Header(None)):
    _check_admin(x_admin_password)
    error_data = []
    log_dir = "logs"
    if os.path.exists(log_dir):
        for session in os.listdir(log_dir):
            session_dir = os.path.join(log_dir, session)
            if not os.path.isdir(session_dir):
                continue
            for filename in os.listdir(session_dir):
                if not (filename.startswith("errors_") and filename.endswith(".json")):
                    continue
                try:
                    with open(os.path.join(session_dir, filename), 'r', encoding='utf-8') as f:
                        errors = json.load(f)
                    error_data.extend(errors)
                except Exception:
                    pass
    error_data.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
    return {"errors": error_data[:100], "total": len(error_data)}

# ─── 관리자 엔드포인트 ───────────────────────────────────────

@app.get("/api/admin/stats")
async def admin_stats(x_admin_password: str = Header(None)):
    _check_admin(x_admin_password)
    stats = Logger.get_global_stats()
    stats["cache_size"] = len(_cache)
    stats["cache_hits"] = _cache_hits
    return stats

@app.post("/api/admin/upload")
async def admin_upload(
    file: UploadFile = File(...),
    x_admin_password: str = Header(None)
):
    _check_admin(x_admin_password)
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="PDF 파일만 업로드 가능합니다.")
    os.makedirs("data", exist_ok=True)
    dest = os.path.join("data", file.filename)
    content = await file.read()
    with open(dest, "wb") as f:
        f.write(content)
    return {"message": f"{file.filename} 업로드 완료"}

@app.post("/api/admin/rebuild")
async def admin_rebuild(
    background_tasks: BackgroundTasks,
    x_admin_password: str = Header(None)
):
    _check_admin(x_admin_password)
    if rebuild_status["running"]:
        raise HTTPException(status_code=409, detail="이미 재구축 중입니다.")

    def _rebuild():
        global rebuild_status
        rebuild_status = {"running": True, "message": "PDF 로딩 중..."}
        try:
            processors = load_all_pdfs_from_folder("data")
            rebuild_status["message"] = f"{len(processors)}개 PDF 청킹 중..."
            all_chunks = []
            for p in processors:
                all_chunks.extend(qa_engine.vector_store.create_chunks_from_pdf(p))
            rebuild_status["message"] = f"{len(all_chunks)}개 청크 임베딩 중..."
            qa_engine.vector_store.create_vectorstore(all_chunks)
            qa_engine._build_bm25_index()
            _cache.clear()  # 캐시 초기화
            rebuild_status = {"running": False, "message": f"완료 ({len(all_chunks)}개 청크)"}
        except Exception as e:
            rebuild_status = {"running": False, "message": f"오류: {str(e)}"}

    background_tasks.add_task(_rebuild)
    return {"message": "DB 재구축 시작"}

@app.get("/api/admin/rebuild/status")
async def admin_rebuild_status(x_admin_password: str = Header(None)):
    _check_admin(x_admin_password)
    return rebuild_status

@app.post("/api/admin/backup")
async def admin_backup(x_admin_password: str = Header(None)):
    _check_admin(x_admin_password)
    _backup_vectordb()
    return {"message": "백업 완료"}

@app.post("/api/admin/cache/clear")
async def admin_cache_clear(x_admin_password: str = Header(None)):
    _check_admin(x_admin_password)
    count = len(_cache)
    _cache.clear()
    return {"message": f"캐시 {count}개 삭제 완료"}

@app.get("/api/admin/logs")
async def admin_logs(x_admin_password: str = Header(None)):
    _check_admin(x_admin_password)
    log_dir = "logs"
    if not os.path.exists(log_dir):
        return {"sessions": []}
    sessions = [d for d in os.listdir(log_dir) if os.path.isdir(os.path.join(log_dir, d))]
    result = []
    for session in sessions:
        logger = Logger(session_id=session)
        dates = logger.get_all_log_dates()
        result.append({"session_id": session, "dates": dates})
    return {"sessions": result}

# ─── 설계 파라미터 산출 엔드포인트 ─────────────────────────────

class CalcRequest(BaseModel):
    road_grade:   str
    design_speed: int
    terrain:      str
    region:       str
    lane_count:   int

@app.post("/api/calc")
async def calc_params(body: CalcRequest):
    if body.road_grade not in VALID_SPEEDS:
        raise HTTPException(status_code=400, detail="유효하지 않은 도로 등급입니다.")
    valid = VALID_SPEEDS[body.road_grade]
    if body.design_speed not in valid:
        raise HTTPException(
            status_code=400,
            detail=f"{body.road_grade}의 유효한 설계속도: {valid} km/h"
        )
    result = calculate_all(
        body.road_grade, body.design_speed,
        body.terrain, body.region, body.lane_count
    )
    return result

# ── 검증 엔드포인트 ──────────────────────────────────────────

_VALIDATE_LIMIT = {
    "차로폭":                      "min",
    "우측 길어깨폭":               "min",
    "좌측 길어깨폭":               "min",
    "중앙분리대폭 (최소)":         "min",
    "노면 횡단경사":               "fixed",
    "길어깨 횡단경사":             "fixed",
    "최소 평면곡선반경":           "min",
    "최대 편경사":                 "max",
    "완화곡선 최소 길이":          "min",
    "확폭량 (최소 반경 기준)":     "min",
    "최대 종단경사":               "max",
    "최소 종단경사":               "min",
    "볼록형 종단곡선 K값":         "min",
    "오목형 종단곡선 K값":         "min",
    "정지시거":                    "min",
    "앞지르기시거":                "min",
}

class ValidateRequest(BaseModel):
    road_grade:    str
    design_speed:  int
    terrain:       str
    region:        str
    lane_count:    int
    design_values: dict  # {항목명: float}

@app.post("/api/calc/validate")
async def validate_params(body: ValidateRequest):
    result = calculate_all(
        body.road_grade, body.design_speed,
        body.terrain, body.region, body.lane_count
    )

    items = []
    pass_cnt = fail_cnt = na_cnt = skip_cnt = 0

    for p in result["params"]:
        name     = p["name"]
        standard = p["value"]
        is_na    = (standard is None or standard == "해당 없음")
        ltype    = _VALIDATE_LIMIT.get(name, "min")
        dval     = body.design_values.get(name)

        if is_na:
            status  = "na"
            message = "해당 없음"
            na_cnt += 1
        elif dval is None:
            status  = "skip"
            message = "미입력"
            skip_cnt += 1
        elif ltype == "fixed":
            status  = "pass"
            message = f"기준값 그대로 적용 ({standard} {p['unit']})"
            pass_cnt += 1
        elif ltype == "min":
            if dval >= standard:
                status  = "pass"
                message = f"{dval} {p['unit']} ≥ {standard} {p['unit']}"
                pass_cnt += 1
            else:
                status  = "fail"
                message = (
                    f"{dval} {p['unit']} < 기준 최솟값 {standard} {p['unit']}"
                    f" — {p['source']}"
                )
                fail_cnt += 1
        else:  # max
            if dval <= standard:
                status  = "pass"
                message = f"{dval} {p['unit']} ≤ {standard} {p['unit']}"
                pass_cnt += 1
            else:
                status  = "fail"
                message = (
                    f"{dval} {p['unit']} > 기준 최댓값 {standard} {p['unit']}"
                    f" — {p['source']}"
                )
                fail_cnt += 1

        items.append({
            "name":       name,
            "category":   p["category"],
            "standard":   standard,
            "unit":       p["unit"],
            "design":     dval,
            "limit_type": ltype,
            "status":     status,
            "message":    message,
            "source":     p["source"],
        })

    return {
        "items": items,
        "summary": {
            "pass": pass_cnt,
            "fail": fail_cnt,
            "na":   na_cnt,
            "skip": skip_cnt,
        },
    }

@app.get("/api/calc/excel")
async def calc_excel(
    road_grade:   str = Query(...),
    design_speed: int = Query(...),
    terrain:      str = Query(...),
    region:       str = Query(...),
    lane_count:   int = Query(...),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl이 설치되어 있지 않습니다.")

    result    = calculate_all(road_grade, design_speed, terrain, region, lane_count)
    today_str = datetime.now().strftime("%Y년 %m월 %d일")
    inputs    = result["inputs"]

    wb = openpyxl.Workbook()

    # ── 공통 스타일 ──────────────────────────────────────────────────────
    DARK_BLUE  = "1E3A5F"
    BLUE       = "2563EB"
    LIGHT_BLUE = "DBEAFE"
    BG_BLUE    = "EFF6FF"
    GREEN      = "DCFCE7"
    GRAY       = "F1F5F9"

    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor=BLUE)
    sec_fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
    cat_fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
    pass_fill = PatternFill("solid", fgColor=GREEN)
    na_fill   = PatternFill("solid", fgColor=GRAY)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_va   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def hdr_row(ws, row_num, num_cols):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row_num, c)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = thin

    def bc(ws, row_num, col_start, col_end):
        """border + center 일괄 적용"""
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row_num, c)
            cell.border = thin; cell.alignment = center

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 1: 프로젝트 개요
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws1 = wb.active
    ws1.title = "프로젝트 개요"

    ws1.append(["도로 설계 파라미터 산출 보고서", "", ""])         # row 1
    ws1.merge_cells("A1:C1")
    ws1["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws1["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws1["A1"].alignment = center
    ws1.row_dimensions[1].height = 32

    ws1.append(["도로의 구조·시설 기준에 관한 규칙 / 도로설계기준 기반 자동 산출", "", ""])  # row 2
    ws1.merge_cells("A2:C2")
    ws1["A2"].font = Font(size=9, italic=True, color="666666")
    ws1["A2"].fill = PatternFill("solid", fgColor=BG_BLUE)
    ws1["A2"].alignment = center
    ws1.row_dimensions[2].height = 16

    ws1.append(["", "", ""])                                        # row 3 (빈 행)

    ws1.append(["■ 작성 정보", "", ""])                            # row 4
    ws1.merge_cells("A4:C4")
    ws1["A4"].font = Font(bold=True, size=11)
    ws1["A4"].fill = sec_fill
    ws1["A4"].alignment = left_va
    ws1.row_dimensions[4].height = 20

    ws1.append(["작성일", today_str, ""])                          # row 5
    ws1.merge_cells("B5:C5")
    ws1["A5"].font = Font(bold=True)
    bc(ws1, 5, 1, 3)

    ws1.append(["적용 기준", "도로의 구조·시설 기준에 관한 규칙 / 도로설계기준", ""])  # row 6
    ws1.merge_cells("B6:C6")
    ws1["A6"].font = Font(bold=True)
    bc(ws1, 6, 1, 3)

    ws1.append(["", "", ""])                                        # row 7

    ws1.append(["■ 설계 조건 입력", "", ""])                       # row 8
    ws1.merge_cells("A8:C8")
    ws1["A8"].font = Font(bold=True, size=11)
    ws1["A8"].fill = sec_fill
    ws1["A8"].alignment = left_va
    ws1.row_dimensions[8].height = 20

    ws1.append(["항목", "선택값", "비고"])                         # row 9
    hdr_row(ws1, 9, 3)

    for label, val, note in [
        ("도로 등급",  str(inputs["road_grade"]),        "고속도로·일반국도·지방도·시군도"),
        ("설계속도",   f"{inputs['design_speed']} km/h", ""),
        ("지형 조건",  str(inputs["terrain"]),           "평지·구릉지·산지"),
        ("지역 조건",  str(inputs["region"]),            "지방지역·도시지역"),
        ("차로 수",    f"{inputs['lane_count']}차로",    "전체 차로 수"),
    ]:
        ws1.append([label, val, note])
        r = ws1.max_row
        ws1.cell(r, 1).font = Font(bold=True)
        bc(ws1, r, 1, 3)

    ws1.append([""])
    ws1.append(["※ 본 보고서는 참고용 자동 산출값입니다. 실제 설계 적용 시 전문가 검토가 필요합니다.", "", ""])
    r = ws1.max_row
    ws1.merge_cells(f"A{r}:C{r}")
    ws1.cell(r, 1).font = Font(size=9, italic=True, color="888888")
    ws1.cell(r, 1).alignment = left_va

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 28

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 2: 설계 기준값 산출표
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws2 = wb.create_sheet("설계 기준값 산출표")
    ws2.append(["구분", "항목", "기준값", "단위", "적용 조건", "근거 조항", "산출 공식"])
    hdr_row(ws2, 1, 7)

    prev_cat = None
    for p in result["params"]:
        ws2.append([
            p["category"], p["name"], p["value"], p["unit"],
            p["condition"], p["source"], p.get("formula", ""),
        ])
        r = ws2.max_row
        is_new_cat = (p["category"] != prev_cat)
        for col in range(1, 8):
            cell = ws2.cell(r, col)
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if is_new_cat:
                cell.fill = cat_fill
        if is_new_cat:
            prev_cat = p["category"]

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 30
    ws2.column_dimensions["F"].width = 26
    ws2.column_dimensions["G"].width = 46

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 3: 기준 적합성 검토표
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws3 = wb.create_sheet("기준 적합성 검토표")
    ws3.append(["구분", "항목", "기준값", "단위", "설계 적용값", "적합 여부", "검토 기준"])
    hdr_row(ws3, 1, 7)

    LIMIT_TYPE = {
        "차로폭":                      ("min", "설계값 ≥ 최솟값"),
        "우측 길어깨폭":               ("min", "설계값 ≥ 최솟값"),
        "좌측 길어깨폭":               ("min", "설계값 ≥ 최솟값"),
        "중앙분리대폭 (최소)":         ("min", "설계값 ≥ 최솟값"),
        "노면 횡단경사":               ("fixed", "기준값 그대로 적용"),
        "길어깨 횡단경사":             ("fixed", "기준값 그대로 적용"),
        "최소 평면곡선반경":           ("min", "설계값 ≥ 최솟값"),
        "최대 편경사":                 ("max", "설계값 ≤ 최댓값"),
        "완화곡선 최소 길이":          ("min", "설계값 ≥ 최솟값"),
        "확폭량 (최소 반경 기준)":     ("min", "설계값 ≥ 최솟값"),
        "최대 종단경사":               ("max", "설계값 ≤ 최댓값"),
        "최소 종단경사":               ("min", "설계값 ≥ 최솟값"),
        "볼록형 종단곡선 K값":         ("min", "설계값 ≥ 최솟값"),
        "오목형 종단곡선 K값":         ("min", "설계값 ≥ 최솟값"),
        "정지시거":                    ("min", "설계값 ≥ 최솟값"),
        "앞지르기시거":                ("min", "설계값 ≥ 최솟값 (2차로 전용)"),
    }

    prev_cat3 = None
    for p in result["params"]:
        val   = p["value"]
        name  = p["name"]
        is_na = (val is None or val == "해당 없음")
        _, check_desc = LIMIT_TYPE.get(name, ("min", "설계값 ≥ 최솟값"))

        if is_na:
            conform_str = "- 해당없음"
            design_val  = "-"
        else:
            conform_str = "✅ 충족"
            design_val  = val

        ws3.append([p["category"], name, val, p["unit"], design_val, conform_str, check_desc])
        r = ws3.max_row
        is_new_cat = (p["category"] != prev_cat3)
        for col in range(1, 8):
            cell = ws3.cell(r, col)
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col == 6:
                cell.fill = pass_fill if not is_na else na_fill
            elif is_new_cat:
                cell.fill = cat_fill
        if is_new_cat:
            prev_cat3 = p["category"]

    ws3.append([""])
    ws3.append(["※ '설계 적용값' 열에 실제 설계값을 입력하고 기준값과 비교하여 충족 여부를 검토하세요.", "", "", "", "", "", ""])
    r = ws3.max_row
    ws3.merge_cells(f"A{r}:G{r}")
    ws3.cell(r, 1).font      = Font(size=9, italic=True, color="888888")
    ws3.cell(r, 1).alignment = left_va

    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 8
    ws3.column_dimensions["E"].width = 16
    ws3.column_dimensions["F"].width = 14
    ws3.column_dimensions["G"].width = 28

    # ── 직렬화 & 반환 ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"도로설계보고서_{road_grade}_{design_speed}kmh.xlsx"
    encoded  = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )

# ─── 일반 엔드포인트 ─────────────────────────────────────────

@app.get("/api/pdf/{filename}")
async def get_pdf(filename: str):
    pdf_path = os.path.join("data", filename)
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail=f"파일을 찾을 수 없습니다: {filename}")
    encoded_filename = quote(filename)
    return FileResponse(
        pdf_path,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{encoded_filename}"}
    )

@app.get("/api/documents")
async def get_documents():
    data_folder = "data"
    if not os.path.exists(data_folder):
        return {"documents": []}
    pdf_files = [f for f in os.listdir(data_folder) if f.lower().endswith('.pdf')]
    return {"documents": pdf_files}

@app.delete("/api/admin/documents/{filename}")
async def admin_delete_document(filename: str, x_admin_password: str = Header(None)):
    _check_admin(x_admin_password)
    pdf_path = os.path.join("data", filename)
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")
    os.remove(pdf_path)
    qa_engine.vector_store.delete_by_source(filename)
    qa_engine._build_bm25_index()
    _cache.clear()
    return {"message": f"{filename} 삭제 완료"}

@app.get("/api/logs/{session_id}")
async def get_logs(session_id: str, date: str = None):
    try:
        logger = Logger(session_id=session_id)
        logs = logger.get_logs(date) if date else logger.get_logs()
        return {"session_id": session_id, "logs": logs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/logs/{session_id}/dates")
async def get_log_dates(session_id: str):
    try:
        logger = Logger(session_id=session_id)
        dates = logger.get_all_log_dates()
        return {"session_id": session_id, "dates": dates}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/health")
async def health():
    return {"status": "ok"}

# ─── 프론트엔드 서빙 ─────────────────────────────────────────

FRONTEND_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend", "dist")
if os.path.exists(FRONTEND_DIR):
    app.mount("/assets", StaticFiles(directory=os.path.join(FRONTEND_DIR, "assets")), name="static")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        index_path = os.path.join(FRONTEND_DIR, "index.html")
        return HTMLResponse(open(index_path, encoding="utf-8").read())
